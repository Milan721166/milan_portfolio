from flask import Flask, render_template, request, redirect, url_for
from openpyxl import load_workbook, Workbook
import os

app = Flask(__name__)


EXCEL_FILE = "contact_data.xlsx"

@app.route("/")
def home():
    return render_template("index.html")

@app.route("/submit", methods=["POST"])
def submit():
    
    name = request.form.get("name")
    email = request.form.get("email")
    message = request.form.get("message")
    
    
    if not os.path.exists(EXCEL_FILE):
        
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Name"
        sheet["B1"] = "Email"
        sheet["C1"] = "Message"
    else:
        
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active
    
    
    next_row = sheet.max_row + 1
    
    
    sheet[f"A{next_row}"] = name
    sheet[f"B{next_row}"] = email
    sheet[f"C{next_row}"] = message
    
    
    workbook.save(EXCEL_FILE)
    
    return redirect(url_for("thank_you"))

@app.route("/thank-you")
def thank_you():
    return render_template("thank_you.html")


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
